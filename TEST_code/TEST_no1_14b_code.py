import os
try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("❌ openpyxl 라이브러리가 없어요! 터미널에 'pip install openpyxl'을 입력하세요.")
    exit()

file_path = 'D:/KIVOSY_LOG.xlsx'

# 1. 파일이 있는지 확인하고 없으면 새로 만듭니다
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "KIVOSY_LOG"
    print(f"✨ 새 파일을 생성합니다: {file_path}")
else:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    print(f"📂 기존 파일을 불러옵니다: {file_path}")

# 2. 내용 기록 (A1 셀에 환경 개선 필요 적기)
ws['A1'] = '환경 개선 필요'
ws['B1'] = '업데이트 시간: ' + str(os.path.getmtime(file_path) if os.path.exists(file_path) else "방금 전")

# 3. 저장
try:
    wb.save(file_path)
    print("✅ 엑셀 파일 저장 완료! 이제 D드라이브를 보세요!")
except PermissionError:
    print("❌ 엑셀 파일이 이미 열려있어서 저장할 수 없어요! 엑셀을 끄고 다시 하세요.")